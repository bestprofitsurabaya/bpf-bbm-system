"""
Unit Tests untuk generate_trip_logsheet (modules/excel_generator.py)
BPF BBM System

Jalankan:
    docker exec bbm_web python3 -m pytest tests/test_excel_generator.py -v
"""
import sys
import os
from io import BytesIO

# Tambahkan parent directory ke path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from modules.excel_generator import generate_trip_logsheet


def _sample_master():
    return {
        'trip_date': '01/08/2026',
        'driver_name': 'TEST DRIVER',
        'nopol': 'L 1234 ABC',
        'jam_keberangkatan': '08:00',
        'jam_tiba': '17:00',
        'km_awal': 1000,
        'km_akhir': 1120,
    }


def _sample_details():
    return [
        {'lokasi_berangkat': 'Kantor', 'pukul_berangkat': '08:00', 'km_berangkat': 1000,
         'lokasi_tujuan': 'Surabaya', 'pukul_tujuan': '09:00', 'km_tujuan': 1060},
        {'lokasi_berangkat': 'Surabaya', 'pukul_berangkat': '10:00', 'km_berangkat': 1060,
         'lokasi_tujuan': 'Gresik', 'pukul_tujuan': '11:00', 'km_tujuan': 1120},
    ]


class TestTripLogsheet:
    def test_returns_xlsx_bytes(self):
        """Harus mengembalikan bytes file .xlsx (magic ZIP)."""
        result = generate_trip_logsheet(_sample_master(), _sample_details())
        assert isinstance(result, bytes)
        assert len(result) > 1000, "Output xlsx terlalu kecil"
        assert result[:2] == b'PK', "Bukan file xlsx valid (magic PK)"

    def test_empty_details_still_works(self):
        """Tanpa rute detail pun harus tetap menghasilkan file valid."""
        result = generate_trip_logsheet(_sample_master(), [])
        assert isinstance(result, bytes)
        assert result[:2] == b'PK'

    def test_total_km_calculated(self):
        """Total jarak harus terhitung: (1060-1000)+(1120-1060) = 120 km."""
        from openpyxl import load_workbook
        result = generate_trip_logsheet(_sample_master(), _sample_details())
        wb = load_workbook(BytesIO(result))
        ws = wb.active
        # total_row = 11 + len(details) + 1 = 14; E14 = total KM
        assert ws['E14'].value == '120 KM', f"Total KM salah: {ws['E14'].value!r}"

    def test_header_fields_present(self):
        """Kop surat (tanggal, driver, plat) harus terisi benar."""
        from openpyxl import load_workbook
        result = generate_trip_logsheet(_sample_master(), _sample_details())
        wb = load_workbook(BytesIO(result))
        ws = wb.active
        assert ws['B3'].value == 'Tanggal: 01/08/2026'
        assert ws['E3'].value == 'DRIVER: TEST DRIVER'
        assert ws['G3'].value == 'PLAT: L 1234 ABC'
