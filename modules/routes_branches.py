"""API Cabang (v2.19.2 — multi-cabang).

Dikelola Admin: daftar cabang, tambah/edit (identitas + nama database),
aktif/nonaktif, buat database cabang (salinan skema), ganti cabang aktif
(sesi) + laporan konsolidasi lintas cabang (PDF/Excel, v2.21) —
memungkinkan satu instalasi melayani banyak cabang.
"""
from datetime import datetime
from flask import request, jsonify, session, make_response

from modules.config import get_master_connection, get_db_connection
from modules.helpers import role_required, log_activity_async
from modules import branch_manager as bm
from modules.security import rate_limit

_DUMMY_TX_SQL = """INSERT IGNORE INTO transactions
    (driver_name, nopol, vehicle_type, bbm_type, nominal, liter, price_per_liter,
     odo_km, spbu_type, status, km_per_liter, jumlah_appointment, is_dummy, gps_address)
    VALUES
    ('AKHAD','L 1413 CBI','AVANZA','PERTALITE',200000,20.00,10000,12936,'rekanan','archived',12.50,3,1,'Jl. Raya Darmo 45, Surabaya'),
    ('AHMAT','B 2628 SRP','INNOVA','PERTAMAX',270000,20.00,13500,71126,'rekanan','archived',10.20,5,1,'Jl. Ahmad Yani 120, Surabaya')"""


def register_branch_routes(app):

    @app.route('/api/branches')
    @role_required(['admin'])
    def api_branches_list():
        try:
            return jsonify({'branches': bm.list_branches()})
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    @app.route('/api/branches/stats')
    @role_required(['admin'])
    def api_branches_stats():
        """Statistik ringkas per cabang untuk dashboard Admin."""
        try:
            return jsonify({'branches': bm.branch_stats()})
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    @app.route('/api/branches/current')
    def api_branches_current():
        """Info cabang sesi + daftar cabang yang bisa dipilih (admin: semua aktif)."""
        try:
            current = bm.current_branch()
            branches = []
            if session.get('user_role') == 'admin':
                branches = [{
                    'code': b['code'], 'name': b['name'],
                    'db_name': b['db_name'], 'is_active': bool(b['is_active']),
                } for b in bm.list_branches() if b.get('is_active')]
            else:
                branches = [{'code': current['code'], 'name': current['name'],
                             'db_name': current['db_name'], 'is_active': True}]
            return jsonify({'current': current, 'branches': branches})
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    @app.route('/api/branches/save', methods=['POST'])
    @role_required(['admin'])
    def api_branches_save():
        try:
            data = request.get_json(silent=True) or {}
            branch, err = bm.save_branch(data)
            if err:
                return jsonify({'status': 'error', 'msg': err}), 400
            actor = session.get('full_name') or session.get('user_name') or 'Admin'
            log_activity_async(0, 'branch_save', 'admin', actor,
                               new_data={'code': branch['code'], 'db_name': branch['db_name']},
                               ip=request.remote_addr)
            # Bila diminta & db_name baru → langsung buat database cabang
            msg = f'Cabang {branch["code"]} disimpan'
            if data.get('ensure_db') and branch.get('db_name'):
                ok, m = bm.ensure_branch_database(branch['code'])
                msg = f'{msg}. {m}'
                if not ok:
                    return jsonify({'status': 'error', 'msg': msg}), 400
            return jsonify({'status': 'success', 'msg': msg, 'branch': branch})
        except Exception as e:
            return jsonify({'status': 'error', 'msg': str(e)}), 500

    @app.route('/api/branches/<code>/seed-demo', methods=['POST'])
    @role_required(['admin'])
    @rate_limit(limit=10, window=60, scope='seed-demo')
    def api_branches_seed_demo(code):
        """Tanam data demo (rute + transaksi dummy) langsung ke DB cabang tertentu.

        Berguna untuk gladi resik cabang baru tanpa harus ganti cabang aktif
        — idempoten (display_id DEMO-* dilewati).
        """
        try:
            from scripts.seed_demo_routes import seed_demo_appointments
            branch = bm.get_branch(code)
            if not branch or not branch.get('is_active'):
                return jsonify({'status': 'error', 'msg': 'Cabang tidak ditemukan / nonaktif'}), 404
            conn = get_db_connection(branch_code=code)
            if not conn:
                return jsonify({'status': 'error', 'msg': 'DB cabang tidak tersedia — klik 🗄️ DB dulu'}), 400
            res = seed_demo_appointments(conn=conn, commit=False)
            if res['error']:
                conn.close()
                return jsonify({'status': 'error', 'msg': res['error']}), 400
            cursor = conn.cursor()
            cursor.execute(_DUMMY_TX_SQL)
            dummy_tx = cursor.rowcount
            conn.commit()
            cursor.close()
            conn.close()
            actor = session.get('full_name') or session.get('user_name') or 'Admin'
            log_activity_async(0, 'branch_seed_demo', 'admin', actor,
                               new_data={'code': code, 'routes': res['created'],
                                         'skipped': res['skipped'], 'transactions': dummy_tx},
                               ip=request.remote_addr, branch_code=code)
            return jsonify({'status': 'success',
                            'msg': f'Data demo cabang {code}: {res["created"]} rute baru, {dummy_tx} transaksi dummy'})
        except Exception as e:
            return jsonify({'status': 'error', 'msg': str(e)}), 500

    @app.route('/api/branches/report-pdf')
    @role_required(['admin'])
    def api_branches_report_pdf():
        """PDF ringkasan per cabang (transaksi, kunjungan hari ini, user)."""
        try:
            from modules.pdf_generator import BranchSummaryPDF
            stats = bm.branch_stats()
            pdf = BranchSummaryPDF()
            pdf.add_page()
            pdf.generate(stats)
            raw = pdf.output()
            pdf_bytes = raw.encode('latin-1') if isinstance(raw, str) else bytes(raw)
            response = make_response(pdf_bytes)
            response.headers['Content-Type'] = 'application/pdf'
            response.headers['Content-Disposition'] = 'attachment; filename=BPF_Ringkasan_Cabang.pdf'
            return response
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    @app.route('/api/branches/<code>/activate', methods=['POST'])
    @role_required(['admin'])
    def api_branches_activate(code):
        try:
            ok = bm.set_branch_active(code, True)
            if not ok:
                return jsonify({'status': 'error', 'msg': 'Cabang tidak ditemukan'}), 404
            return jsonify({'status': 'success', 'msg': f'Cabang {code} aktif'})
        except Exception as e:
            return jsonify({'status': 'error', 'msg': str(e)}), 500

    @app.route('/api/branches/<code>/deactivate', methods=['POST'])
    @role_required(['admin'])
    def api_branches_deactivate(code):
        try:
            ok = bm.set_branch_active(code, False)
            if not ok:
                return jsonify({'status': 'error', 'msg': 'Cabang tidak ditemukan'}), 404
            # Jangan biarkan sesi admin tertinggal di cabang nonaktif
            if session.get('branch_code') == code:
                session['branch_code'] = None
                session['branch_name'] = None
            return jsonify({'status': 'success', 'msg': f'Cabang {code} nonaktif'})
        except Exception as e:
            return jsonify({'status': 'error', 'msg': str(e)}), 500

    @app.route('/api/branches/<code>/ensure-db', methods=['POST'])
    @role_required(['admin'])
    def api_branches_ensure_db(code):
        try:
            ok, msg = bm.ensure_branch_database(code)
            if not ok:
                return jsonify({'status': 'error', 'msg': msg}), 400
            return jsonify({'status': 'success', 'msg': msg})
        except Exception as e:
            return jsonify({'status': 'error', 'msg': str(e)}), 500

    @app.route('/api/branches/switch', methods=['POST'])
    @role_required(['admin'])
    def api_branches_switch():
        """Ganti cabang aktif di sesi (Admin bisa mengoperasikan cabang mana pun)."""
        try:
            data = request.get_json(silent=True) or {}
            code = str(data.get('code', '') or '').strip().upper()
            branch = bm.get_branch(code)
            if not branch or not branch.get('is_active'):
                return jsonify({'status': 'error', 'msg': 'Cabang tidak ditemukan / nonaktif'}), 404
            session['branch_code'] = branch['code']
            session['branch_name'] = branch['name']
            log_activity_async(0, 'branch_switch', 'admin',
                               session.get('full_name') or session.get('user_name') or 'Admin',
                               new_data={'code': branch['code']}, ip=request.remote_addr)
            return jsonify({'status': 'success', 'msg': f'Cabang aktif: {branch["name"]}',
                            'current': bm.current_branch()})
        except Exception as e:
            return jsonify({'status': 'error', 'msg': str(e)}), 500

    # ================================================================
    # KONSOLIDASI LINTAS CABANG (v2.21) — PDF / Excel / JSON
    # ================================================================
    def _consolidated_payload():
        """Statistik semua cabang + transaksi BBM terbaru dari tiap DB."""
        stats = bm.branch_stats()
        latest = {}
        for b in stats:
            code = b['code']
            db = b.get('db_name') or ''
            if not b.get('is_active') or not db:
                latest[code] = []
                continue
            try:
                from modules.config import _pool_for, DB_CONFIG
                if db == DB_CONFIG['database']:
                    c = get_master_connection()
                else:
                    pool = _pool_for(db)
                    c = pool.get_connection() if pool else None
                if not c:
                    latest[code] = []
                    continue
                cur = c.cursor(dictionary=True)
                cur.execute("SELECT display_id, driver_name, nopol, bbm_type, nominal, liter, created_at "
                            "FROM transactions WHERE status='archived' ORDER BY created_at DESC LIMIT 5")
                rows = cur.fetchall()
                cur.close(); c.close()
                latest[code] = rows
            except Exception:
                latest[code] = []
        return stats, latest

    @app.route('/api/branches/consolidated')
    @role_required(['admin'])
    def api_branches_consolidated():
        stats, latest = _consolidated_payload()
        return jsonify({'branches': stats, 'latest_by_branch': latest})

    @app.route('/api/branches/consolidated-pdf')
    @role_required(['admin'])
    def api_branches_consolidated_pdf():
        from modules.pdf_generator import ConsolidatedReportPDF
        stats, latest = _consolidated_payload()
        pdf = ConsolidatedReportPDF()
        pdf.add_page()
        pdf.generate(stats, latest, generated_by=session.get('full_name') or session.get('user_name') or 'Admin')
        from io import BytesIO
        buf = BytesIO()
        pdf.output(buf)
        buf.seek(0)
        resp = make_response(buf.getvalue())
        resp.headers['Content-Type'] = 'application/pdf'
        resp.headers['Content-Disposition'] = f'attachment; filename=konsolidasi_cabang_{datetime.now().strftime("%Y%m%d")}.pdf'
        log_activity_async(0, 'branch_consolidated_pdf', 'admin',
                           session.get('full_name') or session.get('user_name') or 'Admin',
                           new_data={'count': len(stats)}, ip=request.remote_addr)
        return resp

    @app.route('/api/branches/consolidated-excel')
    @role_required(['admin'])
    def api_branches_consolidated_excel():
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        stats, latest = _consolidated_payload()

        wb = Workbook()
        ws = wb.active
        ws.title = 'Ringkasan'
        ws.append(['KODE', 'CABANG', 'DATABASE', 'TRANSAKSI', 'KUNJUNGAN HARI INI', 'USER', 'STATUS'])
        hdr = PatternFill('solid', fgColor='1F2937')
        for c in ws[1]:
            c.font = Font(bold=True, color='FFFFFF'); c.fill = hdr; c.alignment = Alignment(horizontal='center')
        for b in stats:
            ws.append([b.get('code'), b.get('name'), b.get('db_name'), b.get('transactions', 0),
                       b.get('appointments_today', 0), b.get('users', 0),
                       'Aktif' if b.get('is_active') else 'Nonaktif'])
        ws2 = wb.create_sheet('Transaksi Terbaru')
        ws2.append(['CABANG', 'ID', 'DRIVER', 'NOPOL', 'BBM', 'NOMINAL', 'LITER', 'TANGGAL'])
        for c in ws2[1]:
            c.font = Font(bold=True, color='FFFFFF'); c.fill = hdr; c.alignment = Alignment(horizontal='center')
        for code, txs in latest.items():
            for t in (txs or [])[:5]:
                ws2.append([code, t.get('display_id'), t.get('driver_name'), t.get('nopol'),
                            t.get('bbm_type'), float(t.get('nominal') or 0), float(t.get('liter') or 0),
                            t.get('created_at').strftime('%d-%m-%Y %H:%M') if hasattr(t.get('created_at'), 'strftime') else t.get('created_at')])
        for sheet in (ws, ws2):
            for col in sheet.columns:
                letter = col[0].column_letter
                sheet.column_dimensions[letter].width = max(12, max((len(str(c.value or '')) for c in col), default=8) + 2)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = make_response(buf.getvalue())
        resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        resp.headers['Content-Disposition'] = f'attachment; filename=konsolidasi_cabang_{datetime.now().strftime("%Y%m%d")}.xlsx'
        log_activity_async(0, 'branch_consolidated_excel', 'admin',
                           session.get('full_name') or session.get('user_name') or 'Admin',
                           new_data={'count': len(stats)}, ip=request.remote_addr)
        return resp
