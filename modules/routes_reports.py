"""Report, Rekap, Analytics & Settings Routes"""
from flask import (render_template, request, redirect, url_for, flash, jsonify, make_response)
from modules.config import get_db_connection
from modules.helpers import log_activity_async
from modules.engine import get_rekap_data
from modules.pdf_generator import PDFReportCompact, BBMReportPDF
from datetime import datetime, timedelta
import zipfile, io, os, subprocess
from io import BytesIO

def register_report_routes(app):
    
    @app.route('/admin/report/<int:tx_id>')
    def generate_report(tx_id):
        try:
            conn = get_db_connection()
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT * FROM transactions WHERE id=%s", (tx_id,))
            tx = cursor.fetchone()
            if not tx:
                return "Data tidak ditemukan", 404
            cursor.execute("UPDATE transactions SET is_reported=TRUE WHERE id=%s", (tx_id,))
            conn.commit()
            cursor.close(); conn.close()
            pdf = PDFReportCompact()
            pdf.add_page()
            pdf.generate_compact_report(tx, app.config['UPLOAD_FOLDER'])
            pdf_raw = pdf.output(dest='S')
            pdf_bytes = pdf_raw.encode('latin-1') if isinstance(pdf_raw, str) else bytes(pdf_raw)
            response = make_response(pdf_bytes)
            response.headers['Content-Type'] = 'application/pdf'
            response.headers['Content-Disposition'] = f'attachment; filename=Report_BBM_{tx["nopol"]}_{tx["id"]}.pdf'
            return response
        except Exception as e:
            return f"Error: {str(e)}", 500

    @app.route('/admin/rekap')
    def admin_rekap():
        try:
            page = request.args.get('page', 1, type=int)
            per_page = 75
            today = datetime.now().date()
            week_ago = (datetime.now() - timedelta(days=7)).date()
            sd = request.args.get('start_date', '').strip()
            ed = request.args.get('end_date', '').strip()
            if not sd and not ed and not request.args.get('nopol') and not request.args.get('driver'):
                sq, eq = week_ago.strftime('%Y-%m-%d'), today.strftime('%Y-%m-%d')
                default = True
            else:
                sq = sd if sd else None
                eq = ed if ed else None
                default = False
            filters = {'start_date': sd if sd else (week_ago.strftime('%Y-%m-%d') if default else ''),
                       'end_date': ed if ed else (today.strftime('%Y-%m-%d') if default else ''),
                       'nopol': request.args.get('nopol', ''), 'driver': request.args.get('driver', '')}
            all_txs = get_rekap_data(start_date=sq, end_date=eq,
                                     nopol=filters['nopol'] if filters['nopol'] else None,
                                     driver=filters['driver'] if filters['driver'] else None)
            total = len(all_txs)
            pages = max((total+per_page-1)//per_page, 1)
            if page<1: page=1
            if page>pages: page=pages
            paginated = all_txs[(page-1)*per_page:page*per_page]
            pagination = {'page': page, 'total_pages': pages, 'total_count': total,
                          'has_prev': page>1, 'has_next': page<pages, 'prev_page': page-1, 'next_page': page+1}
            return render_template('rekap.html', transactions=paginated, filters=filters, pagination=pagination)
        except Exception as e:
            return f"Error: {str(e)}", 500

    @app.route('/admin/rekap/pdf')
    def rekap_pdf():
        try:
            filters = {'start_date': request.args.get('start_date','').strip(), 'end_date': request.args.get('end_date','').strip(),
                       'nopol': request.args.get('nopol','').strip(), 'driver': request.args.get('driver','').strip()}
            data = get_rekap_data(start_date=filters['start_date'] if filters['start_date'] else None,
                                  end_date=filters['end_date'] if filters['end_date'] else None,
                                  nopol=filters['nopol'] if filters['nopol'] else None,
                                  driver=filters['driver'] if filters['driver'] else None)
            if not data:
                flash('Tidak ada data', 'warning')
                return redirect(url_for('admin_rekap'))
            ts = []
            ts.append(f"DRIVER: {filters['driver'].upper()}" if filters['driver'] else "DRIVER: ALL")
            ts.append(f"NOPOL: {filters['nopol'].upper()}" if filters['nopol'] else "NOPOL: ALL")
            if filters['start_date'] or filters['end_date']:
                s, e = filters['start_date'] or 'ALL', filters['end_date'] or 'ALL'
                ts.append(f"PERIODE: {s} S/D {e}")
            title = "REKAP DANA BBM - " + " | ".join(ts)
            pdf = BBMReportPDF(title=title)
            pdf.add_page()
            pdf.generate_table(data)
            pdf_raw = pdf.output(dest='S')
            pdf_bytes = pdf_raw.encode('latin-1') if isinstance(pdf_raw, str) else bytes(pdf_raw)
            response = make_response(pdf_bytes)
            response.headers['Content-Type'] = 'application/pdf'
            response.headers['Content-Disposition'] = f'inline; filename=Rekap_BBM.pdf'
            return response
        except Exception as e:
            return f"Error: {str(e)}", 500

    @app.route('/admin/analytics')
    def admin_analytics():
        try:
            conn = get_db_connection()
            if not conn: return "DB error", 500
            today = datetime.now().date()
            default_start = (datetime.now() - timedelta(days=30)).date()
            start_date = request.args.get('start_date', default_start.strftime('%Y-%m-%d'))
            end_date = request.args.get('end_date', today.strftime('%Y-%m-%d'))
            cursor = conn.cursor(dictionary=True)
            cursor.execute("""
                SELECT t.nopol, t.vehicle_type, t.bbm_type, COALESCE(AVG(NULLIF(t.km_per_liter,0)), 0) as avg_km_per_liter,
                       COUNT(*) as total_transactions, SUM(t.jumlah_appointment) as jumlah_appointment,
                       SUM(t.nominal) as total_nominal, SUM(t.liter) as total_liter
                FROM transactions t WHERE t.status IN ('verified','archived','os_finance','verified_ga','rejected')
                  AND DATE(t.created_at)>=%s AND DATE(t.created_at)<=%s
                GROUP BY t.nopol, t.vehicle_type, t.bbm_type ORDER BY avg_km_per_liter DESC
            """, (start_date, end_date))
            vehicle_perf = cursor.fetchall() or []
            cursor.execute("""
                SELECT DATE(created_at) as date, AVG(NULLIF(km_per_liter,0)) as avg_efficiency,
                       COUNT(*) as transactions, SUM(nominal) as total_nominal
                FROM transactions WHERE status IN ('verified','archived','os_finance','verified_ga','rejected')
                  AND DATE(created_at)>=%s AND DATE(created_at)<=%s
                GROUP BY DATE(created_at) ORDER BY date DESC
            """, (start_date, end_date))
            daily_stats = cursor.fetchall() or []
            summary = {
                'total_vehicles': len(vehicle_perf) if vehicle_perf else 0,
                'total_tx': sum([v.get('total_transactions', 0) or 0 for v in vehicle_perf]) if vehicle_perf else 0,
                'total_nominal': sum([v.get('total_nominal', 0) or 0 for v in vehicle_perf]) if vehicle_perf else 0,
                'total_liter': round(sum([v.get('total_liter', 0) or 0 for v in vehicle_perf]), 1) if vehicle_perf else 0
            }
            cursor.close(); conn.close()
            return render_template('analytics.html', vehicle_performance=vehicle_perf, daily_stats=daily_stats,
                                  filters={'start_date': start_date, 'end_date': end_date}, summary=summary)
        except Exception as e:
            return f"Error: {str(e)}", 500

    @app.route('/admin/logs')
    def admin_logs_view():
        try:
            conn = get_db_connection()
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT id, transaction_id, action, user_type, user_name, created_at, ip_address FROM activity_logs ORDER BY created_at DESC LIMIT 150")
            logs = cursor.fetchall()
            cursor.close(); conn.close()
            return render_template('logs.html', logs=logs)
        except Exception as e:
            return f"Error: {str(e)}", 500

    @app.route('/admin/riwayat')
    def admin_riwayat():
        return redirect(url_for('admin_rekap'))

    @app.route('/finance/download-archive/<int:tx_id>')
    def download_archive(tx_id):
        try:
            conn = get_db_connection()
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT * FROM transactions WHERE id=%s", (tx_id,))
            tx = cursor.fetchone()
            cursor.close(); conn.close()
            if not tx: return "Not found", 404
            memory_file = BytesIO()
            with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zf:
                for field, label in [('foto_odo_sebelum','01_ODO'),('foto_nota_odo_sesudah','02_Nota'),('foto_struk','03_Struk'),('foto_struk_dispenser','04_Dispenser')]:
                    fn = tx.get(field)
                    if fn:
                        fp = os.path.join(app.config['UPLOAD_FOLDER'], fn)
                        if os.path.exists(fp):
                            zf.write(fp, f"{label}_{tx['nopol']}{os.path.splitext(fn)[1]}")
            memory_file.seek(0)
            response = make_response(memory_file.read())
            response.headers['Content-Type'] = 'application/zip'
            response.headers['Content-Disposition'] = f'attachment; filename=Archive_{tx["nopol"]}_{tx["id"]}.zip'
            return response
        except Exception as e:
            return f"Error: {str(e)}", 500

    @app.route('/admin/backup')
    def backup_database():
        try:
            filename = f'backup_bpf_bbm_{datetime.now().strftime("%Y%m%d_%H%M%S")}.sql'
            log_activity_async(0, 'backup_download', 'admin', 'Admin', new_data={'filename': filename})
            cmd = ['mysqldump', '--ssl=0', '-h', 'db', '-u', 'bpf_user', '-pbpf_pass', 'bpf_asset_system']
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
            if result.returncode != 0:
                return make_response("Backup unavailable", 500)
            response = make_response(result.stdout)
            response.headers['Content-Type'] = 'application/sql'
            response.headers['Content-Disposition'] = f'attachment; filename={filename}'
            return response
        except Exception as e:
            return f"Error: {str(e)}", 500

    @app.route('/admin/templates/import-bbm')
    def download_import_template():
        from openpyxl import Workbook
        wb = Workbook(); ws = wb.active; ws.title = "Template Import BBM"
        ws.append(["Tanggal (DD/MM/YYYY HH:MM)", "Nama Driver", "No Polisi", "Tipe Kendaraan", "Jenis BBM", "Nominal (Rp)", "Harga Per Liter", "Odometer (KM)", "Jumlah Appointment", "Alamat GPS"])
        ws.append(["02/07/2026 14:00", "AKHAD", "L 1413 CBI", "AVANZA", "PERTALITE", 200000, 10000, 12936, 3, "Jl. Raya Darmo 45, Surabaya"])
        out = io.BytesIO(); wb.save(out); out.seek(0)
        log_activity_async(0, 'template_download', 'admin', 'Admin', ip=request.remote_addr)
        response = make_response(out.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = 'attachment; filename=Template_Import_BBM.xlsx'
        return response

    @app.route('/admin/settings/import-xlsx', methods=['POST'])
    def import_xlsx_data():
        from openpyxl import load_workbook
        from modules.helpers import ensure_all_master_data
        if 'excel_file' not in request.files:
            flash('Tidak ada file!', 'error')
            return redirect(url_for('admin_settings'))
        file = request.files['excel_file']
        if file.filename == '':
            flash('Nama file kosong!', 'error')
            return redirect(url_for('admin_settings'))
        try:
            wb = load_workbook(file, data_only=True); ws = wb.active
            conn = get_db_connection(); cursor = conn.cursor()
            success = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[1] or not row[2]: continue
                try:
                    created_at = row[0] if isinstance(row[0], datetime) else datetime.strptime(str(row[0]), "%d/%m/%Y %H:%M")
                except:
                    created_at = datetime.now()
                driver_name = str(row[1]).strip().upper()
                nopol = str(row[2]).strip().upper()
                vehicle_type = str(row[3]).strip().upper() if row[3] else "AVANZA"
                bbm_type = str(row[4]).strip().upper() if row[4] else "PERTALITE"
                nominal = float(row[5] or 0)
                price_per_liter = float(row[6] or 10000)
                liter = nominal/price_per_liter if price_per_liter>0 else 0
                odo_km = int(row[7] or 0)
                jumlah_appt = int(row[8] or 0)
                gps_address = str(row[9]) if row[9] else "Import Data Historis"
                ensure_all_master_data(driver_name, nopol, vehicle_type, bbm_type, price_per_liter)
                cursor.execute("INSERT INTO transactions (driver_name, nopol, vehicle_type, bbm_type, nominal, liter, price_per_liter, odo_km, spbu_type, status, ml_anomaly_flag, km_per_liter, gps_address, jumlah_appointment, created_at) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,'rekanan','archived',0,0,%s,%s,%s)",
                    (driver_name, nopol, vehicle_type, bbm_type, nominal, liter, price_per_liter, odo_km, gps_address, jumlah_appt, created_at))
                success += 1
            conn.commit(); cursor.close(); conn.close()
            log_activity_async(0, 'import_excel', 'admin', 'Admin', new_data={'total_rows': success}, ip=request.remote_addr)
            flash(f'✅ {success} data berhasil diimpor!', 'success')
        except Exception as e:
            flash(f'Error: {str(e)}', 'error')
        return redirect(url_for('admin_settings'))
