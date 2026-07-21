
# ============================================================
# TRIP LOGSHEET EXPORT
# ============================================================
def generate_trip_logsheet(master, details):
    """Generate corporate trip logsheet Excel file with strict layout"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.drawing.image import Image as ExcelImage
    import io, os
    from datetime import datetime

    wb = Workbook()
    ws = wb.active
    ws.title = "Logsheet Harian"

    # --- KOP SURAT: Logo Perusahaan ---
    logo_path = os.path.join('static', 'icon-512.png')
    if os.path.exists(logo_path):
        try:
            img = ExcelImage(logo_path)
            img.width = 55
            img.height = 55
            ws.add_image(img, 'B2')
        except:
            pass

    # Column widths
    widths = {'A': 3, 'B': 28, 'C': 14, 'D': 10, 'E': 28, 'F': 14, 'G': 10, 'H': 15, 'I': 15}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Styles
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    thick_bottom = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='medium')
    )
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    title_font = Font(name='Arial', bold=True, size=14)
    subtitle_font = Font(name='Arial', bold=True, size=10)
    normal_font = Font(name='Arial', size=9)
    small_font = Font(name='Arial', size=8)

    # Row 1: Empty
    # Row 2: Company Title
    ws.merge_cells('B2:I2')
    cell = ws['B2']
    cell.value = 'PT. BESTPROFIT FUTURES - Cab. Surabaya'
    cell.font = title_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

    # Row 3: Date, Driver, Plat
    ws['B3'] = f"Tanggal: {master['trip_date']}"
    ws['B3'].font = subtitle_font
    ws['E3'] = f"DRIVER: {master['driver_name']}"
    ws['E3'].font = subtitle_font
    ws['G3'] = f"PLAT: {master['nopol']}"
    ws['G3'].font = subtitle_font

    # Row 4: "Mengetahui,"
    ws.merge_cells('G4:I4')
    ws['G4'] = 'Mengetahui,'
    ws['G4'].font = subtitle_font
    ws['G4'].alignment = Alignment(horizontal='center')

    # Row 5: Times, Chief Driver, GA&HRD
    ws['B5'] = f"Jam Keberangkatan: {str(master['jam_keberangkatan'])[:5]}"
    ws['B5'].font = normal_font
    ws['E5'] = f"Jam Tiba: {str(master.get('jam_tiba', ''))[:5] if master.get('jam_tiba') else '-'}"
    ws['E5'].font = normal_font
    ws['G5'] = '(Chief Driver)'
    ws['G5'].font = small_font
    ws['G5'].alignment = Alignment(horizontal='center')
    ws['I5'] = '(GA & HRD)'
    ws['I5'].font = small_font
    ws['I5'].alignment = Alignment(horizontal='center')

    # Row 6-7: Empty
    # Row 8: KM Awal, KM Akhir
    ws['B8'] = f"KM Awal: {master['km_awal']:,}"
    ws['B8'].font = subtitle_font
    ws['E8'] = f"KM Akhir: {master.get('km_akhir', 0):,}" if master.get('km_akhir') else "KM Akhir: -"
    ws['E8'].font = subtitle_font

    # Row 9-10: Empty
    # Row 11: Table Headers
    headers = ['No.', 'Lokasi Berangkat', 'Pukul', 'KM', 'Lokasi Tujuan', 'Pukul', 'KM', 'Tanda Tangan', 'Ket.']
    row = 11
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(name='Arial', bold=True, size=9)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = header_fill

    # Row 12+: Dynamic Data
    total_km = 0
    for i, detail in enumerate(details, 1):
        row = 11 + i
        pukul_berangkat = str(detail['pukul_berangkat'])[:5] if detail.get('pukul_berangkat') else '-'
        pukul_tujuan = str(detail['pukul_tujuan'])[:5] if detail.get('pukul_tujuan') else '-'
        km_berangkat = detail.get('km_berangkat', 0) or 0
        km_tujuan = detail.get('km_tujuan', 0) or 0
        jarak = km_tujuan - km_berangkat if km_tujuan > km_berangkat else 0
        total_km += jarak

        values = [
            i,
            detail.get('lokasi_berangkat', '-'),
            pukul_berangkat,
            km_berangkat,
            detail.get('lokasi_tujuan', '-'),
            pukul_tujuan,
            km_tujuan,
            '',
            f'{jarak} km' if jarak > 0 else ''
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = normal_font
            cell.border = thin_border
            if col_idx in [1, 3, 4, 6, 7]:
                cell.alignment = Alignment(horizontal='center')

    # Total row
    total_row = 11 + len(details) + 1
    ws.merge_cells(f'A{total_row}:D{total_row}')
    # ===== ASSIGNMENT HISTORY (Fitur #5) =====
    try:
        import mysql.connector
        conn_assign = mysql.connector.connect(
            host=os.environ.get('DB_HOST', 'db'),
            user=os.environ.get('DB_USER', 'bpf_user'),
            password=os.environ.get('DB_PASSWORD', 'bpf_pass'),
            database=os.environ.get('DB_NAME', 'bpf_asset_system')
        )
        cur = conn_assign.cursor(dictionary=True)
        cur.execute("SELECT * FROM vehicle_assignments WHERE nopol=%s ORDER BY id DESC LIMIT 5", (master['nopol'],))
        assignments = cur.fetchall()
        cur.close()
        conn_assign.close()
        
        if assignments:
            assign_row = total_row + 2
            ws.merge_cells(f'A{assign_row}:I{assign_row}')
            ws[f'A{assign_row}'] = 'RIWAYAT SERAH TERIMA KENDARAAN'
            ws[f'A{assign_row}'].font = Font(name='Arial', bold=True, size=10)
            ws[f'A{assign_row}'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            
            assign_headers = ['Driver', 'Tgl Mulai', 'Tgl Selesai', 'Durasi', 'Dikonfirmasi', 'Catatan']
            for col_idx, h in enumerate(assign_headers, 1):
                cell = ws.cell(row=assign_row+1, column=col_idx, value=h)
                cell.font = Font(bold=True, size=9)
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            for i, a in enumerate(assignments):
                r = assign_row + 2 + i
                ws.cell(row=r, column=1, value=a['driver_name']).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                ws.cell(row=r, column=2, value=str(a['assigned_date'])).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                ws.cell(row=r, column=3, value=str(a.get('unassigned_date', '-'))).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                ws.cell(row=r, column=4, value=f"{a.get('duration_days', '-')} hari").border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                ws.cell(row=r, column=5, value='✅' if a.get('confirmed_by_driver') else '❌').border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                ws.cell(row=r, column=6, value=a.get('driver_notes', '-')).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    except:
        pass
    
    # ===== ASSIGNMENT HISTORY (Fitur #5) =====
    try:
        import mysql.connector
        conn_assign = mysql.connector.connect(
            host=os.environ.get('DB_HOST', 'db'),
            user=os.environ.get('DB_USER', 'bpf_user'),
            password=os.environ.get('DB_PASSWORD', 'bpf_pass'),
            database=os.environ.get('DB_NAME', 'bpf_asset_system')
        )
        cur = conn_assign.cursor(dictionary=True)
        cur.execute("SELECT * FROM vehicle_assignments WHERE nopol=%s ORDER BY id DESC LIMIT 5", (master['nopol'],))
        assignments = cur.fetchall()
        cur.close()
        conn_assign.close()
        
        if assignments:
            assign_row = total_row + 2
            ws.merge_cells(f'A{assign_row}:I{assign_row}')
            ws[f'A{assign_row}'] = 'RIWAYAT SERAH TERIMA KENDARAAN'
            ws[f'A{assign_row}'].font = Font(name='Arial', bold=True, size=10)
            ws[f'A{assign_row}'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            
            assign_headers = ['Driver', 'Tgl Mulai', 'Tgl Selesai', 'Durasi', 'Dikonfirmasi', 'Catatan']
            for col_idx, h in enumerate(assign_headers, 1):
                cell = ws.cell(row=assign_row+1, column=col_idx, value=h)
                cell.font = Font(bold=True, size=9)
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            for i, a in enumerate(assignments):
                r = assign_row + 2 + i
                ws.cell(row=r, column=1, value=a['driver_name']).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                ws.cell(row=r, column=2, value=str(a['assigned_date'])).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                ws.cell(row=r, column=3, value=str(a.get('unassigned_date', '-'))).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                ws.cell(row=r, column=4, value=f"{a.get('duration_days', '-')} hari").border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                ws.cell(row=r, column=5, value='✅' if a.get('confirmed_by_driver') else '❌').border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                ws.cell(row=r, column=6, value=a.get('driver_notes', '-')).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    except:
        pass
    
    ws[f'A{total_row}'] = 'TOTAL JARAK TEMPUH'
    ws[f'A{total_row}'].font = Font(name='Arial', bold=True, size=10)
    ws[f'A{total_row}'].border = thick_bottom
    ws[f'E{total_row}'] = f'{total_km:,} KM'
    ws[f'E{total_row}'].font = Font(name='Arial', bold=True, size=10)
    ws[f'E{total_row}'].border = thick_bottom

    # Signature rows (after 2 empty rows)
    sig_row = total_row + 3
    ws[f'G{sig_row}'] = f"Surabaya, {datetime.now().strftime('%d %B %Y')}"
    ws[f'G{sig_row}'].font = normal_font
    ws[f'G{sig_row}'].alignment = Alignment(horizontal='center')
    ws[f'G{sig_row+3}'] = f"( {master['driver_name']} )"
    ws[f'G{sig_row+3}'].font = subtitle_font
    ws[f'G{sig_row+3}'].alignment = Alignment(horizontal='center')

    # Print settings
    ws.sheet_properties.pageSetUpPr = None
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
